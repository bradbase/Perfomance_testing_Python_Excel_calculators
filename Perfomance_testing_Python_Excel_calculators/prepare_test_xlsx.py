
from random import random
from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

# init Excel_individual_SUM.xlsx
dest_filename = 'Nested_sum'
max_rows = 10001

beginning = datetime.now()
print("creating the formulae")
wb = Workbook()
ws1 = wb.active
ws1.title = "Sheet1"
ws1["A1"] = "Value 1"
ws1["B1"] = "Value 2"
ws1["C1"] = "Value 3"
ws1["D1"] = "SUM 1"
ws1["E1"] = "SUM 2"
ws1["F1"] = "SUM 3"
ws1["G1"] = "SUM 4"
ws1["H1"] = "SUM 5"
ws1["I1"] = "SUM 6"
ws1["J1"] = "SUM 7"
ws1["K1"] = "SUM 8"
ws1["L1"] = "SUM 9"
ws1["M1"] = "SUM 10"
ws1["N1"] = "SUM 11"
ws1["O1"] = "SUM 12"
ws1["P1"] = "SUM 13"
ws1["Q1"] = "SUM 14"
ws1["R1"] = "SUM 15"
ws1["S1"] = "SUM 16"
ws1["T1"] = "SUM 17"
ws1["U1"] = "SUM 18"
ws1["V1"] = "SUM 19"
ws1["W1"] = "SUM 20"
ws1["X1"] = "SUM 21"
ws1["Y1"] = "SUM 22"
ws1["Z1"] = "SUM 23"

ten_thousand_beginnning = datetime.now()
for row in range(2, max_rows):
    ws1["D{}".format(row)] = "=SUM(A{}, A{}, B{}:C{}, 10)".format(row, row, row, row)
    ws1["E{}".format(row)] = "=SUM(A{}:D{})".format(row, row)
    ws1["F{}".format(row)] = "=SUM(A{}:E{})".format(row, row)
    ws1["G{}".format(row)] = "=SUM(A{}:F{})".format(row, row)
    ws1["H{}".format(row)] = "=SUM(A{}:G{})".format(row, row)
    ws1["I{}".format(row)] = "=SUM(A{}:H{})".format(row, row)
    ws1["J{}".format(row)] = "=SUM(A{}:I{})".format(row, row)
    ws1["K{}".format(row)] = "=SUM(A{}:J{})".format(row, row)
    ws1["L{}".format(row)] = "=SUM(A{}:K{})".format(row, row)
    ws1["M{}".format(row)] = "=SUM(A{}:L{})".format(row, row)
    ws1["N{}".format(row)] = "=SUM(A{}:M{})".format(row, row)
    ws1["O{}".format(row)] = "=SUM(A{}:N{})".format(row, row)
    ws1["P{}".format(row)] = "=SUM(A{}:O{})".format(row, row)
    ws1["Q{}".format(row)] = "=SUM(A{}:P{})".format(row, row)
    ws1["R{}".format(row)] = "=SUM(A{}:Q{})".format(row, row)
    ws1["S{}".format(row)] = "=SUM(A{}:R{})".format(row, row)
    ws1["T{}".format(row)] = "=SUM(A{}:S{})".format(row, row)
    ws1["U{}".format(row)] = "=SUM(A{}:T{})".format(row, row)
    ws1["V{}".format(row)] = "=SUM(A{}:U{})".format(row, row)
    ws1["W{}".format(row)] = "=SUM(A{}:V{})".format(row, row)
    ws1["X{}".format(row)] = "=SUM(A{}:W{})".format(row, row)
    ws1["Y{}".format(row)] = "=SUM(A{}:X{})".format(row, row)
    ws1["Z{}".format(row)] = "=SUM(A{}:Y{})".format(row, row)
    if row % 10000 == 0:
        print("creaing formulae for row {} has taken {} {}".format(row, datetime.now() - ten_thousand_beginnning, datetime.now() - beginning))
        ten_thousand_beginnning = datetime.now()

wb.save(filename = "{}_template.xlsx".format( dest_filename ))
print("total elapsed {}".format(datetime.now() - beginning))


second_beginnig = datetime.now()
ten_thousand_beginnning = datetime.now()
print("generating numbers")
print("loading workbook")
wb = load_workbook(filename = "{}_template.xlsx".format(dest_filename) )
print("loading took {}".format(datetime.now() - second_beginnig))
sheet_ranges = wb['Sheet1']

for row in range(2, max_rows):
    sheet_ranges["A{}".format(row)] = random()
    sheet_ranges["B{}".format(row)] = random()
    sheet_ranges["C{}".format(row)] = random()
    if row % 10000 == 0:
        print("creaing formulae for row {} has taken {} {}".format(row, datetime.now() - ten_thousand_beginnning, datetime.now() - second_beginnig))
        ten_thousand_beginnning = datetime.now()

print("saving")
wb.save(filename = "{}.xlsx".format(dest_filename))
print("total elapsed {}".format(datetime.now() - second_beginnig))
